"""xlsx 產生器與一鍵產檔（含 PDF）測試。"""

from pathlib import Path

import openpyxl
import pytest

from salary_note.calc import compute_payroll
from salary_note.models import Payroll, Period, Teacher
from salary_note.pdf import find_soffice
from salary_note.roster import write_roster
from salary_note.service import generate_files
from salary_note.statement import deduction_text, write_statements

from conftest import SAMPLE_0808


def test_roster_layout_0808(payroll_0808: Payroll, tmp_path: Path):
    calc = compute_payroll(payroll_0808)
    path = write_roster(calc, tmp_path / "r.xlsx")
    ws = openpyxl.load_workbook(path).active

    assert ws.title == "115.08"
    assert ws["A1"].value == "嘉義市立嘉義國民中學115年08月份外籍英語教師薪資印領清冊"
    assert "A1:Q1" in {str(r) for r in ws.merged_cells.ranges}
    # 不足月表頭註記
    assert ws["C3"].value == "本月\n薪資\n16/31"
    assert ws["H3"].value == "勞保\n機補\n16/30"
    assert ws["L3"].value == "預扣\n稅額\n5%"
    assert ws["I3"].value == "健保\n機補"
    # 資料列與公式
    assert ws["A4"].value == "Sample Teacher"
    assert ws["B4"].value == 8
    assert (ws["C4"].value, ws["D4"].value, ws["E4"].value, ws["F4"].value) == (41677, 2581, 516, 0)
    assert ws["G4"].value == "=C4+D4+E4-F4"
    assert ws["K4"].value == "=SUM(G4:J4)"
    assert ws["L4"].value == "=ROUNDDOWN((C4+D4)*0.05,0)"
    assert ws["O4"].value == "=SUM(L4:N4)"
    assert ws["P4"].value == "=G4-O4"
    assert ws["Q4"].value == "聘期\n115/8/16-116/7/15"
    # 健康檢查費用列
    assert (ws["K5"].value, ws["P5"].value, ws["Q5"].value) == (1936, 1936, "健康檢查費用")
    # 空一列、總計、大寫
    assert ws["A7"].value == "總計" and ws["K7"].value == "=SUM(K4:K6)" and ws["P7"].value == "=SUM(P4:P6)"
    assert ws["A8"].value == "總計" and ws["B8"].value == "伍萬伍仟玖佰參拾陸" and ws["P8"].value == "元整"
    assert "B8:O8" in {str(r) for r in ws.merged_cells.ranges}
    # 核章欄
    assert (ws["A12"].value, ws["F12"].value, ws["K12"].value, ws["Q12"].value) == ("教學組長", "出納組長", "會計室", "校長")
    assert (ws["A14"].value, ws["F14"].value) == ("教務主任", "總務主任")
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_area.endswith("$A$1:$Q$14")


def test_roster_full_month_headers(payroll_0809: Payroll, tmp_path: Path):
    ws = openpyxl.load_workbook(write_roster(compute_payroll(payroll_0809), tmp_path / "r.xlsx")).active
    assert ws["C3"].value == "本月\n薪資"
    assert ws["M3"].value == "勞保\n自付"
    # 沒有健康檢查費用 → 第 5 列空、第 6 列總計
    assert ws["K5"].value is None
    assert ws["A6"].value == "總計" and ws["K6"].value == "=SUM(K4:K5)"
    assert ws["B7"].value == "壹拾萬零參佰肆拾"


def test_roster_two_teachers(tmp_path: Path):
    p = Payroll(
        period=Period(roc_year=115, month=9),
        teachers=[
            Teacher(name="A Teacher", salary=1000, health_check=500),
            Teacher(name="B Teacher", salary=2000),
        ],
    )
    ws = openpyxl.load_workbook(write_roster(compute_payroll(p), tmp_path / "r.xlsx")).active
    assert ws["A4"].value == "A Teacher"
    assert ws["Q5"].value == "健康檢查費用"
    assert ws["A6"].value == "B Teacher"
    assert ws["A8"].value == "總計" and ws["K8"].value == "=SUM(K4:K7)"
    assert ws["B9"].value == "參仟伍佰"


def test_statement_0808(payroll_0808: Payroll, tmp_path: Path):
    calc = compute_payroll(payroll_0808)
    wb = openpyxl.load_workbook(write_statements(calc, tmp_path / "s.xlsx"))
    assert wb.sheetnames == ["Sample Teacher"]
    ws = wb.active
    assert ws["A1"].value == "嘉義市嘉義國中外籍教師鐘點通知單\nCYJH Salary Statement"
    labels = [ws[f"A{r}"].value for r in range(2, 13)]
    values = [ws[f"B{r}"].value for r in range(2, 13)]
    assert labels[0] == "姓名\nName" and values[0] == "Sample Teacher"
    assert values[1] == "August 16-31, 2026"
    assert values[2] == "Academic Affairs Office" and values[3] == "Teacher"
    assert values[4:8] == [41677, 2581, 516, 44774]
    assert values[8] == "616 (labor insurance)\n1,359 (health insurance)\n2,212 (withholding tax)\n= 4,187"
    assert labels[9].startswith("健康檢查補助") and values[9] == 1936
    assert labels[10] == "實發金額\nNet Total" and values[10] == 42523
    assert ws.page_setup.orientation == "portrait"


def test_statement_without_health_check(payroll_0809: Payroll, tmp_path: Path):
    ws = openpyxl.load_workbook(write_statements(compute_payroll(payroll_0809), tmp_path / "s.xlsx")).active
    labels = [ws[f"A{r}"].value for r in range(2, 20) if ws[f"A{r}"].value]
    assert not any(l.startswith("健康檢查") for l in labels)
    assert labels[-1] == "實發金額\nNet Total"
    assert ws[f"B{len(labels) + 1}"].value == 79959


def test_statement_sheet_titles_unique(tmp_path: Path):
    p = Payroll(period=Period(roc_year=115, month=9),
                teachers=[Teacher(name="Same/Name", salary=1), Teacher(name="Same/Name", salary=2)])
    wb = openpyxl.load_workbook(write_statements(compute_payroll(p), tmp_path / "s.xlsx"))
    assert wb.sheetnames == ["Same Name", "Same Name (2)"]


def test_deduction_text_uses_thousand_separators(payroll_0809: Payroll):
    t = compute_payroll(payroll_0809).teachers[0]
    assert deduction_text(t) == "1,145 (labor insurance)\n1,359 (health insurance)\n4,287 (withholding tax)\n= 6,791"


def test_generate_files_without_pdf(payroll_0808: Payroll, tmp_path: Path):
    r = generate_files(payroll_0808, tmp_path, make_pdf=False, job="20260901-000000_11508")
    assert r.folder == tmp_path / "20260901-000000_11508"
    names = {f.name: f.kind for f in r.files}
    assert names == {
        "外師薪資印領清冊_115.08.xlsx": "roster-xlsx",
        "外師鐘點通知單_115.08.xlsx": "statement-xlsx",
        "外師薪資_115.08_全部.zip": "zip",
    }
    assert not r.warnings


@pytest.mark.skipif(not find_soffice(), reason="LibreOffice 未安裝")
def test_generate_files_with_pdf(payroll_0808: Payroll, tmp_path: Path):
    import pymupdf

    r = generate_files(payroll_0808, tmp_path, make_pdf=True, job="20260901-000001_11508")
    assert not r.warnings, r.warnings
    kinds = {f.kind for f in r.files}
    assert {"roster-pdf", "statement-pdf", "zip"} <= kinds

    roster_pdf = pymupdf.open(r.folder / "外師薪資印領清冊_115.08.pdf")
    assert roster_pdf.page_count == 1
    page = roster_pdf[0]
    assert page.rect.width > page.rect.height  # 橫式
    text = page.get_text()
    # LibreOffice 有把公式算出來
    for expected in ("115年08月份", "44,774", "54,000", "2,212", "4,187", "40,587", "55,936", "42,523", "伍萬伍仟玖佰參拾陸"):
        assert expected in text, f"缺 {expected}"

    stmt_pdf = pymupdf.open(r.folder / "外師鐘點通知單_115.08.pdf")
    assert stmt_pdf.page_count == 1
    stext = stmt_pdf[0].get_text()
    for expected in ("CYJH Salary Statement", "August 16-31, 2026", "44,774", "= 4,187", "42,523"):
        assert expected in stext, f"缺 {expected}"


def test_sample_dict_is_consistent_with_sheet():
    """conftest 樣本本身就要能還原原表數字。"""
    t = Teacher(**SAMPLE_0808)
    assert t.salary + t.housing + t.transport == 44774
