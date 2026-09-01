"""靜態版（site/，GitHub Pages）測試：以 Python 版為對照答案。

- calc.js 的每個數字要跟 Python compute_payroll 一樣
- 瀏覽器用 ExcelJS 產的 xlsx，用 openpyxl 讀回來跟 Python 版逐格比對（值、公式、字型、對齊、框線、合併、欄寬、列高、頁面設定）
- UI：自動記憶、下載、匯出／匯入、列印檢視
"""

from __future__ import annotations

import base64
import http.server
import io
import json
import threading
from pathlib import Path

import openpyxl
import pytest

playwright = pytest.importorskip("playwright.sync_api")
from playwright.sync_api import sync_playwright  # noqa: E402

from conftest import SAMPLE_0808, SAMPLE_0809  # noqa: E402
from salary_note.calc import compute_payroll  # noqa: E402
from salary_note.models import Payroll  # noqa: E402
from salary_note.roster import build_roster  # noqa: E402
from salary_note.statement import build_statements  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
SITE = ROOT / "site"


class _Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *a, **kw):
        super().__init__(*a, directory=str(SITE), **kw)

    def log_message(self, *a):  # 安靜
        pass


@pytest.fixture(scope="module")
def site_url():
    httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), _Handler)
    threading.Thread(target=httpd.serve_forever, daemon=True).start()
    try:
        yield f"http://127.0.0.1:{httpd.server_address[1]}/"
    finally:
        httpd.shutdown()


@pytest.fixture(scope="module")
def browser():
    with sync_playwright() as p:
        try:
            b = p.chromium.launch()
        except Exception as e:
            pytest.skip(f"Chromium 無法啟動：{e}")
        yield b
        b.close()


@pytest.fixture
def page(browser, site_url):
    ctx = browser.new_context(accept_downloads=True)
    pg = ctx.new_page()
    errors: list[str] = []
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.on("console", lambda m: errors.append(m.text) if m.type == "error" else None)
    pg.on("dialog", lambda d: d.accept())
    pg.errors = errors  # type: ignore[attr-defined]
    pg.goto(site_url)
    pg.wait_for_function("typeof SalaryXlsx !== 'undefined'")
    yield pg
    ctx.close()


def _payload(sample: dict, period: dict, **extra) -> dict:
    return {"period": period, "teachers": [dict(sample)], **extra}


PAYLOADS = {
    "0808": _payload(SAMPLE_0808, {"roc_year": 115, "month": 8, "start_day": 16}),
    "0809": _payload(SAMPLE_0809, {"roc_year": 115, "month": 9}),
    "prorate": _payload(dict(SAMPLE_0809, labor_ins_self=1155), {"roc_year": 115, "month": 8, "start_day": 16}, prorate=True),
    "two": {
        "period": {"roc_year": 115, "month": 9},
        "teachers": [dict(SAMPLE_0809, name="A Teacher", health_check=500),
                     dict(SAMPLE_0809, name="B Teacher", tax_rate_pct=18)],
    },
}

JS_CALC = """(p) => {
  const c = SalaryCalc.computePayroll(p);
  return {
    total_gross: c.total_gross, total_net: c.total_net, upper: c.total_gross_upper,
    common: c.common_tax_rate_pct, roster_title: c.roster_title,
    period: { label_zh: c.period.label_zh, label_en: c.period.label_en, ratio: c.period.ratio_label,
              ins_ratio: c.period.insurance_ratio_label, days: c.period.days, partial: c.period.is_partial },
    teachers: c.teachers.map((t) => ({
      salary: t.salary, housing: t.housing, transport: t.transport,
      labor_ins_employer: t.labor_ins_employer, pension_employer: t.pension_employer, labor_ins_self: t.labor_ins_self,
      subtotal: t.subtotal, gross: t.gross, tax: t.tax, deductions: t.deductions, net: t.net,
      net_with_extras: t.net_with_extras, remark: t.remark,
    })),
  };
}"""


@pytest.mark.parametrize("key", list(PAYLOADS))
def test_calc_js_matches_python(page, key):
    payload = PAYLOADS[key]
    js = page.evaluate(JS_CALC, payload)
    py = compute_payroll(Payroll.model_validate(payload))
    assert js["total_gross"] == py.total_gross
    assert js["total_net"] == py.total_net
    assert js["upper"] == py.total_gross_upper
    assert (js["common"] is None) == (py.common_tax_rate_pct is None)
    assert js["roster_title"] == py.payroll.roster_title
    per = py.period
    assert js["period"] == {"label_zh": per.label_zh, "label_en": per.label_en, "ratio": per.ratio_label,
                            "ins_ratio": per.insurance_ratio_label, "days": per.days, "partial": per.is_partial}
    for jt, pt in zip(js["teachers"], py.teachers, strict=True):
        for k in jt:
            assert jt[k] == getattr(pt, k), f"{key}: {k} js={jt[k]} py={getattr(pt, k)}"
    assert page.errors == []


# ---------- xlsx 逐格比對 ----------

def _js_workbook(page, builder: str, payload: dict) -> openpyxl.Workbook:
    b64 = page.evaluate(
        f"async (p) => SalaryXlsx.toBase64(SalaryXlsx.{builder}(SalaryCalc.computePayroll(p)))", payload)
    return openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))


def _py_workbook(builder, payload: dict) -> openpyxl.Workbook:
    wb = builder(compute_payroll(Payroll.model_validate(payload)))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf)


def _merged(ws) -> set[str]:
    return {str(r) for r in ws.merged_cells.ranges}


def _col_widths(ws) -> dict[str, float]:
    """ExcelJS 會把相同寬度的相鄰欄合寫成 <col min max>，openpyxl 只掛在第一欄；這裡展開成每欄。"""
    from openpyxl.utils import get_column_letter

    out: dict[str, float] = {}
    for dim in ws.column_dimensions.values():
        if dim.width and dim.min and dim.max:
            for i in range(dim.min, dim.max + 1):
                out[get_column_letter(i)] = dim.width
    return out


def _in_merge_slave(ws, cell) -> bool:
    for r in ws.merged_cells.ranges:
        if cell.coordinate in r and (cell.row, cell.column) != (r.min_row, r.min_col):
            return True
    return False


def _assert_sheets_equal(py_ws, js_ws, label: str):
    assert py_ws.title == js_ws.title, label
    assert _merged(py_ws) == _merged(js_ws), label
    for row in py_ws.iter_rows(min_row=1, max_row=py_ws.max_row, max_col=py_ws.max_column):
        for pc in row:
            jc = js_ws[pc.coordinate]
            where = f"{label} {pc.coordinate}"
            assert jc.value == pc.value, f"{where}: js={jc.value!r} py={pc.value!r}"
            if pc.value is None:
                continue
            assert jc.number_format == pc.number_format, f"{where} numFmt js={jc.number_format!r} py={pc.number_format!r}"
            assert (jc.font.name, float(jc.font.size), bool(jc.font.bold)) == (pc.font.name, float(pc.font.size), bool(pc.font.bold)), f"{where} font"
            assert (jc.alignment.horizontal, jc.alignment.vertical, bool(jc.alignment.wrap_text)) == \
                   (pc.alignment.horizontal, pc.alignment.vertical, bool(pc.alignment.wrap_text)), f"{where} align"
            if not _in_merge_slave(py_ws, pc):
                pb, jb = pc.border, jc.border
                assert [bool(getattr(jb, s).style) for s in ("left", "right", "top", "bottom")] == \
                       [bool(getattr(pb, s).style) for s in ("left", "right", "top", "bottom")], f"{where} border"
    pyw, jsw = _col_widths(py_ws), _col_widths(js_ws)
    for col in "ABCDEFGHIJKLMNOPQ":
        pw, jw = pyw.get(col), jsw.get(col)
        if pw or jw:
            assert jw is not None and pw is not None and abs(pw - jw) < 0.05, f"{label} col {col} width js={jw} py={pw}"
    for r in range(1, py_ws.max_row + 1):
        ph, jh = py_ws.row_dimensions[r].height, js_ws.row_dimensions[r].height
        if ph:
            assert jh is not None and abs(jh - ph) < 0.05, f"{label} row {r} height js={jh} py={ph}"
    assert js_ws.page_setup.orientation == py_ws.page_setup.orientation, label
    assert int(js_ws.page_setup.paperSize) == int(py_ws.page_setup.paperSize), label
    assert js_ws.sheet_properties.pageSetUpPr.fitToPage, label
    assert (js_ws.print_area or "").split("!")[-1] == (py_ws.print_area or "").split("!")[-1], label


@pytest.mark.parametrize("key", ["0808", "0809", "two"])
def test_roster_xlsx_matches_python(page, key):
    py_wb = _py_workbook(build_roster, PAYLOADS[key])
    js_wb = _js_workbook(page, "buildRoster", PAYLOADS[key])
    _assert_sheets_equal(py_wb.worksheets[0], js_wb.worksheets[0], f"roster {key}")
    assert page.errors == []


@pytest.mark.parametrize("key", ["0808", "0809", "two"])
def test_statement_xlsx_matches_python(page, key):
    py_wb = _py_workbook(build_statements, PAYLOADS[key])
    js_wb = _js_workbook(page, "buildStatements", PAYLOADS[key])
    assert js_wb.sheetnames == py_wb.sheetnames
    for py_ws, js_ws in zip(py_wb.worksheets, js_wb.worksheets, strict=True):
        _assert_sheets_equal(py_ws, js_ws, f"statement {key} {py_ws.title}")
    assert page.errors == []


# ---------- UI ----------

def _fill(page, payload: dict):
    per = payload["period"]
    page.fill("#roc_year", str(per["roc_year"]))
    page.fill("#month", str(per["month"]))
    page.fill("#start_day", str(per.get("start_day", 1)))
    page.fill("#end_day", "" if per.get("end_day") in (None, "") else str(per["end_day"]))
    blocks = page.locator("#teachers .teacher")
    while blocks.count() < len(payload["teachers"]):
        page.click("#addTeacher")
    for i, t in enumerate(payload["teachers"]):
        block = blocks.nth(i)
        for k, v in t.items():
            if v is not None:
                block.locator(f'[data-k="{k}"]').fill(str(v))
    if payload.get("prorate"):
        page.check("#prorate")


def test_boot_defaults_and_live_preview(page):
    assert page.locator("#teachers .teacher").count() == 1
    assert int(page.input_value("#roc_year")) >= 115
    assert page.input_value("#school_name") == "嘉義市立嘉義國民中學"
    _fill(page, PAYLOADS["0808"])
    preview = page.locator("#teachers .teacher .preview").first.inner_text()
    for expected in ("44,774", "54,000", "2,212", "4,187", "40,587", "42,523"):
        assert expected in preview, preview
    docs = page.locator("#preview").inner_text()
    assert "嘉義市立嘉義國民中學115年08月份外籍英語教師薪資印領清冊" in docs
    assert "伍萬伍仟玖佰參拾陸" in docs
    assert "CYJH Salary Statement" in docs and "August 16-31, 2026" in docs
    assert page.errors == []


def test_autosave_restores_on_reload(page, site_url):
    _fill(page, PAYLOADS["0808"])
    page.wait_for_function("document.getElementById('savedAt').textContent.startsWith('已自動記住')")
    page.goto(site_url)  # 重新打開頁面
    page.wait_for_function("typeof SalaryXlsx !== 'undefined'")
    assert page.locator("#savedAt").inner_text() == "已帶入上次內容"
    assert page.input_value("#start_day") == "16"
    block = page.locator("#teachers .teacher").first
    assert block.locator('[data-k="salary"]').input_value() == "41677"
    assert block.locator('[data-k="name"]').input_value() == "Sample Teacher"
    assert "42,523" in block.locator(".preview").inner_text()


def test_download_export_clear_import(page, tmp_path: Path):
    _fill(page, PAYLOADS["0808"])

    with page.expect_download() as dl:
        page.click("#dlRoster")
    roster_path = tmp_path / dl.value.suggested_filename
    dl.value.save_as(roster_path)
    assert roster_path.name == "外師薪資印領清冊_115.08.xlsx"
    wb = openpyxl.load_workbook(roster_path)
    assert wb.sheetnames == ["115.08", "_salary_note_data"]
    assert wb["_salary_note_data"].sheet_state == "hidden"
    assert wb["115.08"]["K7"].value == "=SUM(K4:K6)"
    embedded = json.loads(wb["_salary_note_data"]["A2"].value)
    assert embedded["payload"]["teachers"][0]["salary"] == 41677

    with page.expect_download() as dl:
        page.click("#dlStatement")
    assert dl.value.suggested_filename == "外師鐘點通知單_115.08.xlsx"

    with page.expect_download() as dl:
        page.click("#exportBtn")
    json_path = tmp_path / dl.value.suggested_filename
    dl.value.save_as(json_path)
    assert json_path.name == "外師薪資設定_115.08.json"
    assert json.loads(json_path.read_text(encoding="utf-8"))["payload"]["period"]["start_day"] == 16

    page.click("#clearBtn")  # confirm 由 fixture 自動接受
    block = page.locator("#teachers .teacher").first
    assert block.locator('[data-k="salary"]').input_value() == "0"
    assert page.input_value("#start_day") == "1"

    page.set_input_files("#importFile", str(roster_path))  # 把上月 Excel 拖回來
    page.wait_for_function("document.getElementById('status').textContent.startsWith('已從')")
    assert block.locator('[data-k="salary"]').input_value() == "41677"
    assert page.input_value("#start_day") == "16"

    page.click("#clearBtn")
    page.set_input_files("#importFile", str(json_path))
    page.wait_for_function("document.getElementById('status').textContent.startsWith('已從')")
    assert block.locator('[data-k="health_check"]').input_value() == "1936"
    assert page.errors == []


def test_print_views(page):
    page.evaluate("() => { window.print = () => { window.__printed = (window.__printed || 0) + 1; }; }")
    _fill(page, PAYLOADS["0808"])
    page.click("#printRoster")
    assert page.evaluate("window.__printed") == 1
    root = page.locator("#print-root")
    assert "55,936" in root.inner_html() and "外籍英語教師薪資印領清冊" in root.inner_html()
    assert "landscape" in page.locator("#page-style").inner_text()
    page.click("#printStatement")
    assert page.evaluate("window.__printed") == 2
    assert "CYJH Salary Statement" in root.inner_html() and "= 4,187" in root.inner_html()
    assert "portrait" in page.locator("#page-style").inner_text()
    assert page.errors == []


def test_card_lines_and_png(page, tmp_path: Path):
    """LINE 圖卡：內容資料正確、每位外師一張、下載的是 1440px 寬的 PNG。"""
    import pymupdf

    rows = page.evaluate("(p) => { const c = SalaryCalc.computePayroll(p); return SalaryCard.lines(c, c.teachers[0]); }", PAYLOADS["0808"])
    by_en = {r["en"]: r for r in rows if r["kind"] != "section"}
    assert by_en["Salary"]["value"] == 41677
    assert by_en["Due amount"]["value"] == 44774
    assert by_en["Labor insurance"]["value"] == -616
    assert by_en["Withholding tax (5%)"]["value"] == -2212
    assert by_en["Deduction"]["value"] == -4187
    assert by_en["Health check reimbursement"] == {"kind": "item", "en": "Health check reimbursement", "zh": "健康檢查補助", "value": 1936, "plus": True}
    assert by_en["Net Total"]["value"] == 42523
    assert [r["en"] for r in rows if r["kind"] == "section"] == ["Earnings", "Deductions", "Other"]
    assert "Leave deduction" not in by_en

    _fill(page, PAYLOADS["0808"])
    assert page.locator("#cardModal").is_hidden()
    page.click("#lineCardBtn")
    page.wait_for_function("!document.getElementById('cardModal').hidden && document.querySelectorAll('#cardList img').length === 1 && document.querySelector('#cardList img').naturalWidth === 1440")
    assert page.locator("#cardList figcaption span").inner_text() == "Sample Teacher"
    with page.expect_download() as dl:
        page.click("#cardList .dlCard")
    png = tmp_path / dl.value.suggested_filename
    dl.value.save_as(png)
    assert png.name == "外師薪資通知圖卡_115.08_Sample Teacher.png"
    assert png.read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"
    pix = pymupdf.Pixmap(str(png))
    assert pix.width == 1440 and pix.height > 1400
    assert page.locator("#cardStatus").inner_text().startswith("已下載")

    page.keyboard.press("Escape")
    assert page.locator("#cardModal").is_hidden()
    assert page.locator("#cardList img").count() == 0

    page.click("#addTeacher")
    page.locator("#teachers .teacher").nth(1).locator('[data-k="name"]').fill("Second Teacher")
    page.click("#lineCardBtn")
    page.wait_for_function("document.querySelectorAll('#cardList img').length === 2")
    page.click("#cardClose")
    assert page.locator("#cardModal").is_hidden()
    assert page.errors == []


def test_card_copy_to_clipboard(page):
    _fill(page, PAYLOADS["0808"])
    page.evaluate("() => { window.__copied = []; navigator.clipboard.write = async (items) => { window.__copied.push(items[0].types); }; }")
    page.click("#lineCardBtn")
    page.wait_for_function("document.querySelectorAll('#cardList img').length === 1")
    page.click("#cardList .copyCard")
    page.wait_for_function("document.getElementById('cardStatus').textContent.startsWith('已複製')")
    assert page.evaluate("window.__copied") == [["image/png"]]
    assert "Ctrl+V" in page.locator("#cardStatus").inner_text()


def test_card_button_requires_name(page):
    page.fill("#roc_year", "115")
    page.fill("#month", "9")
    page.locator("#teachers .teacher").first.locator('[data-k="name"]').fill("")
    page.click("#lineCardBtn")
    page.wait_for_function("document.getElementById('status').textContent.includes('姓名未填')")
    assert page.locator("#cardModal").is_hidden()


def test_missing_name_blocks_download(page):
    page.fill("#roc_year", "115")
    page.fill("#month", "9")
    page.locator("#teachers .teacher").first.locator('[data-k="name"]').fill("")
    page.click("#dlRoster")
    page.wait_for_function("document.getElementById('status').textContent.includes('姓名未填')")
