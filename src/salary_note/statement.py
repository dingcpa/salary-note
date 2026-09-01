"""外籍教師鐘點通知單（CYJH Salary Statement，直式 A4）xlsx 產生器。

版型對照 data/input/通知單樣本_115.08.jpg：兩欄表格，左欄中英文項目名、右欄數值；
每位外師一個工作表（轉 PDF 後一人一頁）。
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from .calc import PayrollCalc, TeacherCalc

KAI = "標楷體"
TNR = "Times New Roman"
NUM = "#,##0"

_THIN = Side(style="thin")
BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_H_TITLE = 60.0
ROW_H_NORMAL = 44.0
ROW_H_TALL = 58.0      # 三行標籤（健康檢查補助）
ROW_H_DEDUCTION = 100.0


def _put(ws: Worksheet, ref: str, value, *, font: Font, fmt: str | None = None):
    c = ws[ref]
    c.value = value
    c.font = font
    c.alignment = CENTER_WRAP
    c.border = BOX
    if fmt:
        c.number_format = fmt
    return c


def bilingual(zh: str, en: str, size: float = 12) -> CellRichText:
    """同一格兩行：中文用標楷體、英文用 Times New Roman（讀回時是純字串 "中文\\n英文"）。"""
    return CellRichText(
        TextBlock(InlineFont(rFont=KAI, sz=size), zh + "\n"),
        TextBlock(InlineFont(rFont=TNR, sz=size), en),
    )


def _sheet_title(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", " ", name).strip()[:28] or "Statement"
    title, n = base, 2
    while title in used:
        title = f"{base[:25]} ({n})"
        n += 1
    used.add(title)
    return title


def deduction_text(t: TeacherCalc) -> str:
    return (
        f"{t.labor_ins_self:,} (labor insurance)\n"
        f"{t.health_ins_self:,} (health insurance)\n"
        f"{t.tax:,} (withholding tax)\n"
        f"= {t.deductions:,}"
    )


Row = tuple[tuple[str, str], object, float, bool]


def statement_rows(calc: PayrollCalc, t: TeacherCalc) -> list[Row]:
    """((中文標籤, 英文標籤), 值, 列高, 是否數字) 的清單；健康檢查補助為 0 時不列。"""
    rows: list[Row] = [
        (("姓名", "Name"), t.name, ROW_H_NORMAL, False),
        (("給薪月份", "payment month"), calc.period.label_en, ROW_H_NORMAL, False),
        (("單位", "Office"), t.teacher.office, ROW_H_NORMAL, False),
        (("職務", "Job title"), t.teacher.job_title, ROW_H_NORMAL, False),
        (("本月薪資", "Salary"), t.salary, ROW_H_NORMAL, True),
        (("住宿津貼", "Housing allowance"), t.housing, ROW_H_NORMAL, True),
        (("交通津貼", "Transportation allowance"), t.transport, ROW_H_NORMAL, True),
        (("應發金額", "Due amount"), t.subtotal, ROW_H_NORMAL, True),
        (("扣款金額", "Deduction"), deduction_text(t), ROW_H_DEDUCTION, False),
    ]
    if t.health_check:
        rows.append((("健康檢查補助", "health check reimbursement"), t.health_check, ROW_H_TALL, True))
    rows.append((("實發金額", "Net Total"), t.net_with_extras, ROW_H_NORMAL, True))
    return rows


def _write_sheet(ws: Worksheet, calc: PayrollCalc, t: TeacherCalc) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34

    ws.row_dimensions[1].height = ROW_H_TITLE
    ws.merge_cells("A1:B1")
    title = bilingual(calc.payroll.statement_title_zh, calc.payroll.statement_title_en, size=14)
    _put(ws, "A1", title, font=Font(name=KAI, size=14))
    ws["B1"].border = BOX

    label_font = Font(name=KAI, size=12)
    value_font = Font(name=TNR, size=12)
    row = 2
    for (zh, en), value, height, is_num in statement_rows(calc, t):
        ws.row_dimensions[row].height = height
        _put(ws, f"A{row}", bilingual(zh, en), font=label_font)
        _put(ws, f"B{row}", value, font=value_font, fmt=NUM if is_num else None)
        row += 1

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:B{row - 1}"


def build_statements(calc: PayrollCalc) -> Workbook:
    wb = Workbook()
    used: set[str] = set()
    for i, t in enumerate(calc.teachers):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = _sheet_title(t.name, used)
        _write_sheet(ws, calc, t)
    return wb


def write_statements(calc: PayrollCalc, path: Path) -> Path:
    wb = build_statements(calc)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
