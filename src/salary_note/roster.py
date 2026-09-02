"""外籍英語教師薪資印領清冊（橫式 A4）xlsx 產生器。

版型對照 data/input 原始 xls 的 115.08 工作表：
  第 1 列標題（A:Q 合併）、第 3 列表頭、第 4 列起每位外師一列（有健康檢查費用者再加一列）、
  空一列、總計（數字）、總計（大寫）、四列後核章欄兩列。
金額欄保留公式，讓學校事後可手動微調。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from .calc import PayrollCalc, TeacherCalc

KAI = "標楷體"
TNR = "Times New Roman"
NUM = "#,##0"

_THIN = Side(style="thin")
BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLS = list("ABCDEFGHIJKLMNOPQ")
LAST_COL = COLS[-1]
COL_WIDTHS = {
    "A": 22.9, "B": 5.9, "C": 8.5, "D": 8.4, "E": 8.4, "F": 6.6, "G": 8.5,
    "H": 7.5, "I": 7.5, "J": 8.4, "K": 9.5, "L": 7.5, "M": 7.5, "N": 8.4,
    "O": 8.9, "P": 10.4, "Q": 13.6,
}

# 表頭（不足月時 C/D/E 加 "16/31"、H/J/M 加 "16/30"，稅率欄加 "5%"）
BASE_HEADERS = {
    "A": "外師姓名", "B": "薪級", "C": "本月\n薪資", "D": "住宿\n津貼", "E": "交通費",
    "F": "請假\n扣薪", "G": "小計", "H": "勞保\n機補", "I": "健保\n機補", "J": "勞退\n機補",
    "K": "應發\n金額", "L": "預扣\n稅額", "M": "勞保\n自付", "N": "健保\n自付",
    "O": "代扣款\n小計", "P": "實領\n金額", "Q": "備註",
}
PAY_COLS = ("C", "D", "E")
INS_COLS = ("H", "J", "M")

ROW_H_TITLE = 39.0
ROW_H_HEADER = 53.45
ROW_H_TEACHER = 52.7
ROW_H_EXTRA = 31.35
SIGNATURES_1 = {"A": "教學組長", "F": "出納組長", "L": "會計室"}
SIGNATURES_2 = {"A": "教務主任", "F": "總務主任", "L": "校長"}


def _put(ws: Worksheet, ref: str, value, *, font: Font, align: Alignment = CENTER,
         fmt: str | None = None, border: Border | None = BOX):
    c = ws[ref]
    c.value = value
    c.font = font
    c.alignment = align
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


def _box_row(ws: Worksheet, row: int):
    for col in COLS:
        ws[f"{col}{row}"].border = BOX


def _headers(calc: PayrollCalc) -> dict[str, str]:
    h = dict(BASE_HEADERS)
    period = calc.period
    if period.is_partial:
        for col in PAY_COLS:
            h[col] += f"\n{period.ratio_label}"
        for col in INS_COLS:
            h[col] += f"\n{period.insurance_ratio_label}"
    rate = calc.common_tax_rate_pct
    if rate is not None:
        h["L"] += f"\n{rate.normalize():f}%"
    return h


def _write_teacher_row(ws: Worksheet, row: int, t: TeacherCalc) -> None:
    ws.row_dimensions[row].height = ROW_H_TEACHER
    _box_row(ws, row)
    num = Font(name=TNR, size=12)
    _put(ws, f"A{row}", t.name, font=num)
    _put(ws, f"B{row}", t.teacher.grade, font=num, fmt=NUM)
    for col, val in (
        ("C", t.salary), ("D", t.housing), ("E", t.transport), ("F", t.leave_deduction),
        ("H", t.labor_ins_employer), ("I", t.health_ins_employer), ("J", t.pension_employer),
        ("M", t.labor_ins_self), ("N", t.health_ins_self),
    ):
        _put(ws, f"{col}{row}", val, font=num, fmt=NUM)
    rate = f"{t.tax_rate.normalize():f}"
    _put(ws, f"G{row}", f"=C{row}+D{row}+E{row}-F{row}", font=num, fmt=NUM)
    _put(ws, f"K{row}", f"=SUM(G{row}:J{row})", font=num, fmt=NUM)
    _put(ws, f"L{row}", f"=ROUNDDOWN((C{row}+D{row})*{rate},0)", font=num, fmt=NUM)
    _put(ws, f"O{row}", f"=SUM(L{row}:N{row})", font=num, fmt=NUM)
    _put(ws, f"P{row}", f"=G{row}-O{row}", font=num, fmt=NUM)
    _put(ws, f"Q{row}", t.remark or None, font=Font(name=KAI, size=10), align=LEFT_WRAP)


def _write_extra_row(ws: Worksheet, row: int, amount: int, label: str) -> None:
    ws.row_dimensions[row].height = ROW_H_EXTRA
    _box_row(ws, row)
    num = Font(name=TNR, size=12)
    _put(ws, f"K{row}", amount, font=num, fmt=NUM)
    _put(ws, f"P{row}", amount, font=num, fmt=NUM)
    _put(ws, f"Q{row}", label, font=Font(name=KAI, size=12), align=LEFT_WRAP)


def build_roster(calc: PayrollCalc) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = calc.period.file_tag

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 標題
    ws.row_dimensions[1].height = ROW_H_TITLE
    ws.merge_cells(f"A1:{LAST_COL}1")
    _put(ws, "A1", calc.payroll.roster_title, font=Font(name=KAI, size=22, bold=True), border=None)
    ws.row_dimensions[2].height = 20.0

    # 表頭
    ws.row_dimensions[3].height = ROW_H_HEADER
    for col, text in _headers(calc).items():
        _put(ws, f"{col}3", text, font=Font(name=KAI, size=12), align=CENTER_WRAP)

    # 資料列
    row = 4
    first_data_row = row
    for t in calc.teachers:
        _write_teacher_row(ws, row, t)
        row += 1
        if t.health_check:
            _write_extra_row(ws, row, t.health_check, "健康檢查費用")
            row += 1
    # 空一列（原表習慣）
    ws.row_dimensions[row].height = ROW_H_EXTRA
    _box_row(ws, row)
    last_data_row = row
    row += 1

    # 總計（數字）
    ws.row_dimensions[row].height = ROW_H_EXTRA
    _box_row(ws, row)
    num = Font(name=TNR, size=12)
    _put(ws, f"A{row}", "總計", font=Font(name=KAI, size=12))
    _put(ws, f"K{row}", f"=SUM(K{first_data_row}:K{last_data_row})", font=num, fmt=NUM)
    _put(ws, f"P{row}", f"=SUM(P{first_data_row}:P{last_data_row})", font=num, fmt=NUM)
    total_row = row
    row += 1

    # 總計（大寫）
    ws.row_dimensions[row].height = ROW_H_EXTRA
    _box_row(ws, row)
    _put(ws, f"A{row}", "總計", font=Font(name=KAI, size=12), align=CENTER_WRAP)
    ws.merge_cells(f"B{row}:O{row}")
    _put(ws, f"B{row}", calc.total_gross_upper, font=Font(name=KAI, size=14))
    _put(ws, f"P{row}", "元整", font=Font(name=KAI, size=12))
    upper_row = row

    # 核章欄
    for offset, h in ((1, 28.5), (2, 24.0), (3, 24.0), (4, 19.5), (5, 49.35), (6, 23.25)):
        ws.row_dimensions[upper_row + offset].height = h
    sig_font = Font(name=KAI, size=14)
    for col, text in SIGNATURES_1.items():
        _put(ws, f"{col}{upper_row + 4}", text, font=sig_font, align=LEFT, border=None)
    for col, text in SIGNATURES_2.items():
        _put(ws, f"{col}{upper_row + 6}", text, font=sig_font, align=LEFT, border=None)
    last_row = upper_row + 6

    # 頁面：橫式 A4、縮放成一頁
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.31, right=0.31, top=0.75, bottom=0.35, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:{LAST_COL}{last_row}"

    # 給測試／除錯用的定位資訊
    ws.sheet_properties.tabColor = None
    wb._roster_rows = {"first_data": first_data_row, "total": total_row, "upper": upper_row}  # type: ignore[attr-defined]
    return wb


def write_roster(calc: PayrollCalc, path: Path) -> Path:
    wb = build_roster(calc)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
